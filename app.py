from flask import Flask, request,render_template,redirect,url_for
from openpyxl import Workbook, load_workbook
import os

app = Flask(__name__)
FILE = "contact_data.xlsx"
@app.route("/")
def home():
    return render_template("index.html")

@app.route("/save", methods=["POST"])
def save():
    name = request.form["name"]
    email = request.form["email"]
    phone = request.form["phone"]
    message = request.form["message"]

    if os.path.exists(FILE):
        wb = load_workbook(FILE)
        sheet = wb.active
    else:
        wb = Workbook()
        sheet = wb.active
        sheet.append(["Name", "Email", "Phone", "Message"])

    sheet.append([name, email, phone, message])
    wb.save(FILE)

    return redirect(url_for("home",success="1"))
 

if __name__ == "__main__":
    app.run()
